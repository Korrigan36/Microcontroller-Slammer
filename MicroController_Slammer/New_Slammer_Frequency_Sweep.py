# -*- coding: utf-8 -*-
"""
Created on Fri Jun 16 09:51:31 2017

@author: v-stpurc
"""
import sys
from PyQt5.QtWidgets import (QApplication, QWidget, QLabel, QRadioButton, QVBoxLayout, QCheckBox, QProgressBar,
    QGroupBox, QComboBox, QLineEdit, QPushButton, QMessageBox, QInputDialog, QDialog, QDialogButtonBox, QSlider, QGridLayout, QHBoxLayout, QFileDialog, QListWidget)
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from PyQt5 import QtCore , QtGui

sys.path.append("C:\\Users\\v-stpur\\Documents\\XBoxScripts\\instrument_Libraries")

import visa
import waveFormLib
import chromaLib
import timerThread
import dutInfoDialog

from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.drawing.image import Image

version_Info = "Python PyQt5 Slammer Frequency Sweep V0.1"

        
class MainWindow(QWidget):

    StartFreq = 0
    StopFreq = 0
    Steps = 0
    logSteps = 0
    
    Scope = 0
    Waveform = 0
    Load = 0
    
    slammer = 0

    duty = 0
    onTime = 0
    amplitude = 0
    configID = 0
    pcbaSN = 0
    productSN = 0
    scopeType = 0
    probeType = 0 
    runNotes = 0 
    vregText = 0 
    vregSelection = 0
    frequencySweep = False
    logSweep = False
    sweepWidth = 0
#    workBook = 0 
#    workSheet = 0 
    vout_DC_Value = 0  
    limitsGood = 0 
    

    sense_Resistor = 0
    
    slamLoad = 0
    chromaLoad = 0
    chromaStartLoad = 0
    chromaStopLoad = 0
    
    scopeShot_Signal = pyqtSignal(str)
    dutyCycle_Signal = pyqtSignal(str, int)
    onTime_Signal = pyqtSignal(str, int)
    slamLoad_Signal = pyqtSignal(str, int)
            
    continueSlider = True
    
    def __init__(self):
        super(MainWindow, self).__init__()
        
        self.initUI()
#        self.openInstruments()
#        self.openFileNameDialog()
 
       
    def initUI(self):
        
        self.setGeometry(300, 300, 600, 200)
        self.setWindowTitle('CFTAC')
        self.setWindowIcon(QIcon('xbox_icon.ico')) 
        
        instrumentGrid = QGridLayout()
        instrumentChoiceGroupBox  = QGroupBox()

        scopeLabel = QLabel(self)
        scopeLabel.setText("Scope")
        instrumentGrid.addWidget(scopeLabel, 0, 0)
  
        self.scopeIDN = QLabel(self)
        self.scopeIDN.setText("Scope")
        instrumentGrid.addWidget(self.scopeIDN, 0, 1)
        
        waveformLabel = QLabel(self)
        waveformLabel.setText("Waveform Generator")
        instrumentGrid.addWidget(waveformLabel, 1, 0)
  
        self.waveformIDN = QLabel(self)
        self.waveformIDN.setText("Waveform")
        instrumentGrid.addWidget(self.waveformIDN, 1, 1)
        
        loadLabel = QLabel(self)
        loadLabel.setText("Chroma Load")
        instrumentGrid.addWidget(loadLabel, 2, 0)
  
        self.loadIDN = QLabel(self)
        self.loadIDN.setText("Chroma")
        instrumentGrid.addWidget(self.loadIDN, 2, 1)
        
        instrumentChoiceGroupBox.setLayout(instrumentGrid)
        
        sliderGrid = QGridLayout()
        sliderGroupBox  = QGroupBox()

        self.sliderLabel = QLabel(self)
        self.sliderLabel.setText("Load Slam")
        sliderGrid.addWidget(self.sliderLabel, 0, 0)

        self.slamLoadSlider = QSlider(Qt.Horizontal, self)
        self.slamLoadSlider.setFocusPolicy(Qt.NoFocus)
        self.slamLoadSlider.setRange(1, 400)
        self.slamLoadSlider.setValue(50)
        self.slamLoadSlider.setGeometry(720, 150, 200, 20)
        self.slamLoadSlider.setEnabled(True)
        self.slamLoadSlider.valueChanged[int].connect(self.slamLoadChanged)

        self.slamLoad = QLabel(self)
        self.slamLoad.setText(str(self.slamLoadSlider.value()) + "amps")
        sliderGrid.addWidget(self.slamLoad, 0, 1)

        sliderGrid.addWidget(self.slamLoadSlider, 1, 0)
        
        self.onTimeLabel = QLabel(self)
        self.onTimeLabel.setText("On Time")
        sliderGrid.addWidget(self.onTimeLabel, 2, 0)

        self.onTimeSlider = QSlider(Qt.Horizontal, self)
        self.onTimeSlider.setFocusPolicy(Qt.NoFocus)
        self.onTimeSlider.setRange(15, 150)
        self.onTimeSlider.setValue(100)
        self.onTimeSlider.setGeometry(720, 150, 200, 20)
        self.onTimeSlider.setEnabled(True)
        self.onTimeSlider.valueChanged[int].connect(self.onTimeChanged)

        self.onTime = QLabel(self)
        self.onTime.setText(str(self.onTimeSlider.value()) + "ns")
        sliderGrid.addWidget(self.onTime, 2, 1)

        sliderGrid.addWidget(self.onTimeSlider, 3, 0)
        
        self.dutyLabel = QLabel(self)
        self.dutyLabel.setText("Duty Cycle")
        sliderGrid.addWidget(self.dutyLabel, 4, 0)

        self.dutyCycleSlider = QSlider(Qt.Horizontal, self)
        self.dutyCycleSlider.setFocusPolicy(Qt.NoFocus)
        self.dutyCycleSlider.setRange(1, 50)
        self.dutyCycleSlider.setValue(50)
        self.dutyCycleSlider.setGeometry(720, 150, 200, 20)
        self.dutyCycleSlider.setEnabled(True)
        self.dutyCycleSlider.valueChanged[int].connect(self.dutyCycleChanged)

        self.dutyCycle = QLabel(self)
        self.dutyCycle.setText(str(self.dutyCycleSlider.value()) + "%")
        sliderGrid.addWidget(self.dutyCycle, 4, 1)

        sliderGrid.addWidget(self.dutyCycleSlider, 5, 0)
        
        sliderGroupBox.setLayout(sliderGrid)
         
        autoTestGroupBox  = QGroupBox()
        autoTestLayout    = QHBoxLayout()
        
        self.microSlam = QCheckBox("Use Micro Slammer")
        self.microSlam.setChecked(True)

        autoTestLayout.addWidget(self.microSlam)
        
        self.autoTest = QCheckBox("Automated Test")
        self.autoTest.setChecked(False)
        self.autoTest.stateChanged.connect(self.autoTestCheck)

        autoTestLayout.addWidget(self.autoTest)
        
        self.tdcLabel = QLabel(self)
        self.tdcLabel.setText("TDC")
        autoTestLayout.addWidget(self.tdcLabel)

        self.tdcList = QLineEdit(self)
        self.tdcList.setEnabled(True)
        autoTestLayout.addWidget(self.tdcList)
        
        self.edcLabel = QLabel(self)
        self.edcLabel.setText("EDC")
        autoTestLayout.addWidget(self.edcLabel)

        self.edcList = QLineEdit(self)
        self.edcList.setEnabled(True)
        autoTestLayout.addWidget(self.edcList)
        
        autoTestGroupBox.setLayout(autoTestLayout)
        
        self.vregComboBox = QComboBox(self)
        self.vregComboBox.addItem("V_3P3_CFX")
        self.vregComboBox.addItem("V_CPUCORE")
        self.vregComboBox.addItem("V_DRAM1P8")
        self.vregComboBox.addItem("V_GFXCORE")
        self.vregComboBox.addItem("V_MEMIO")
        self.vregComboBox.addItem("V_MEMPHY")
        self.vregComboBox.addItem("V_SOC")
        self.vregComboBox.addItem("V_SOC1P8")

        startButtonGroupBox  = QGroupBox()
        startButtonLayout    = QHBoxLayout()
        self.startStopButton = QPushButton('Start Slam', self)
        self.startStopButton.setGeometry(800, 70, 180, 50)

        self.font = QFont()
        self.font.setBold(True)
        self.font.setPointSize(16)
        self.startStopButton.setFont(self.font)
        self.startStopButton.setStyleSheet('QPushButton {background-color: #A3C1DA; color: black;}')
        self.startStopButton.setText("Start Slam")
        self.startStopButton.clicked[bool].connect(self.startStopTest)
        startButtonLayout.addWidget(self.startStopButton)
        startButtonGroupBox.setLayout(startButtonLayout)
 
        quitButtonGroupBox  = QGroupBox()
        quitButtonLayout    = QHBoxLayout()
        self.font.setPointSize(12)
        self.quitButton = QPushButton('Quit', self)
        self.quitButton.setFont(self.font)
        self.quitButton.setGeometry(890, 260, 100, 30)
        self.quitButton.clicked[bool].connect(self.closeEventLocal)
        quitButtonLayout.addWidget(self.quitButton)
        quitButtonGroupBox.setLayout(quitButtonLayout)
        
        grid = QGridLayout()
        grid.setColumnStretch(0,5)
        grid.setColumnStretch(1,5)
        grid.addWidget(instrumentChoiceGroupBox, 0, 0)
        grid.addWidget(sliderGroupBox, 0, 1)
        grid.addWidget(self.vregComboBox, 1, 0)
        grid.addWidget(autoTestGroupBox, 1, 1)
        grid.addWidget(startButtonGroupBox, 2, 0)
        grid.addWidget(quitButtonGroupBox, 2, 1)
        self.setLayout(grid)

        self.show()
        
    def autoTestCheck(self):
        if self.autoTest.isChecked() == True:
            self.dutyCycleSlider.setEnabled(False)
            self.onTimeSlider.setEnabled(False)
            self.slamLoadSlider.setEnabled(False)
            self.tdcList.setEnabled(True)
            self.edcList.setEnabled(True)
        else:
            self.dutyCycleSlider.setEnabled(True)
            self.onTimeSlider.setEnabled(True)
            self.slamLoadSlider.setEnabled(True)
            self.tdcList.setEnabled(False)
            self.edcList.setEnabled(False)
        

    def runLoop(self):
        print ("Got into Run Loop**********************")
        self.continueSlider = True

    def slamLoadChanged(self):
        slamLoad = self.slamLoadSlider.value()        
        self.slamLoad.setText(str(self.slamLoadSlider.value()) + "amps")
        if  self.continueSlider:
            self.continueSlider = False
            self.slamLoad_Signal.emit('Changing slam load to: ' + str(slamLoad), slamLoad)

    def onTimeChanged(self):
        onTime = self.onTimeSlider.value()        
        self.onTime.setText(str(onTime) + "ns")
        if  self.continueSlider:
            self.continueSlider = False
            self.onTime_Signal.emit('Changing onTime to ' + str(onTime) + "us", onTime)

    def dutyCycleChanged(self):
        dutyCycle = self.dutyCycleSlider.value()        
        self.dutyCycle.setText(str(dutyCycle) + "%")
        if  self.continueSlider:
            self.dutyCycle_Signal.emit('Changing duty cycle to ' + str(dutyCycle) + "%", dutyCycle)
            self.continueSlider = False
      
    def closeEventLocal(self, event):
        try:
            self.work.stopTimer()
        except:
            print ("Timer Thread Not Running")
            
        #With openpyxl you name the file when you save it.
        #We don't save it until the exit button is pushed and the gui quits
#        self.workBook.save(self.excelFileName)

        print("Got quit signal")
        self.close()


    def quitLoop(self):
        self.thread.quit()

    def is_number(self, s):
        try:
            float(s)
            return True
        except ValueError:
            return False   
        
    def startStopTest(self):
        returnString = self.startStopButton.text()
        if returnString.find("Start Slam") != -1:
        
            dialog = dutInfoDialog.DutInfoDialog()
            dialog.setWindowModality(Qt.ApplicationModal)
            dialog.exec_()
            self.configID, self.pcbaSN, self.productSN, self.scopeType, self.probeType, self.runNotes, self.sense_Resistor = dialog.returnInfo()
 
            print ("Slam Started")

            self.startStopButton.setStyleSheet('QPushButton {background-color: #FF2000; color: black;}')
            self.startStopButton.font()
            self.startStopButton.setText("Abort Slam")
                
            self.startThread()
        else:
            print ("Slam Aborted")
            self.startStopButton.setStyleSheet('QPushButton {background-color: #A3C1DA; color: black;}')
            self.startStopButton.setText("Start Slam")

            self.thread.quit()
            self.work.stopTimer()

    def startThread(self):

        parameterArray = [self.Scope, self.Load, self.Waveform, self.slamLoadSlider.value(), self.dutyCycleSlider.value(), self.onTimeSlider.value(), self.microSlam.isChecked()]
        self.work = timerThread.TimerThread(parameterArray) 
        self.work.timerSignal.connect(self.runLoop)
        self.work.quitSignal.connect(self.quitLoop)
        self.scopeShot_Signal.connect(self.work.scopeShot_Slot)
        self.dutyCycle_Signal.connect(self.work.dutyCycle_Slot)
        self.onTime_Signal.connect(self.work.onTime_Slot)
        self.slamLoad_Signal.connect(self.work.slamLoad_Slot)
        self.thread = QThread()

        self.work.moveToThread(self.thread)
        self.thread.started.connect(self.work.run)
        self.thread.start()
        
    def openFileNameDialog(self):    

        self.configFileName = QFileDialog.getOpenFileName(self, 'Open file', 'C:\\Users\\v-stpur\\Documents\\XBoxScripts\\SlammerFrequencySweepAllRails',"Excel files (*.xlsx)")
        print ("Filename is: " + str(self.configFileName[0]))

        # Open the Config file
        self.config_workbook = load_workbook(self.configFileName[0], data_only=True)
        self.config_sheet = self.config_workbook.get_sheet_by_name(name = 'AC Matrix') 
        vregName = self.config_sheet.cell(1, 1).value
        print ("first name in config file: " + vregName)


    def openInstruments(self):

        rm = visa.ResourceManager()
        instrumentList = rm.list_resources()
        for loopIndex in range(0, len(instrumentList)):
            inst = rm.open_resource(instrumentList[loopIndex])
                        
            try:
                returnString = inst.query("*IDN?")
                if returnString.find("CHROMA") != -1:
                    print("found chroma")
                    self.str_load_frame_id = returnString
                    self.str_load_frame_id = self.str_load_frame_id.replace (',','_')
                    self.str_load_frame_id = self.str_load_frame_id.replace ('\n','')
                    self.str_load_frame_id = self.str_load_frame_id.replace ('\r','')
                    print (self.str_load_frame_id)
    
                    self.loadIDN.setText(returnString[:14]) 
                    self.Load = inst
                if returnString.find("MSO58") != -1:
                    print("found scope Series 5")
                    self.Scope = inst
    #                self.scopeInstr.setText(instrumentList[loopIndex]) 
                    countComma = 0
                    for i, c in enumerate(returnString):
                        if "," == c:
                            countComma = countComma + 1
    #                        print "comma found in scope string at " + str(i)
                        if countComma == 1:
                            endOfString = i + 1
                    self.scopeIDN.setText(returnString[0:endOfString]) 
                if returnString.find("33621A") != -1:
                    print("found waveform gen")
                    self.Waveform = inst
    #                self.waveformInstr.setText(instrumentList[loopIndex]) 
                    countComma = 0
                    for i, c in enumerate(returnString):
                        if "," == c:
                            countComma = countComma + 1
                        if countComma == 1:
                            endOfString = i + 1
                    self.waveformIDN.setText(returnString[0:endOfString]) 
            except visa.VisaIOError:
                print ("bad juju in instrument list")
     
        tempString = "*RST"
        self.Load.write(tempString) 
      
if __name__ == '__main__':
    
    app = QCoreApplication.instance()
    if app is None:
        app = QApplication(sys.argv)
    ex = MainWindow()
    app.exec_()  
#    sys.exit(app.exec_())  
